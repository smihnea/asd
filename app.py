#!/usr/bin/env python3
"""
Excel Consolidator - Aplicație Avansată de Procesare Fișiere Excel
==================================================================

CONTEXT PENTRU ASISTENȚII LLM:
==============================
Aceasta este o aplicație desktop de producție care consolidează multiple fișiere Excel (.xlsx/.xlsm)
cu structuri de tabele similare într-un workbook unificat cu două foi de output specializate:
1. "Cumulative" - Vedere secvențială a tuturor datelor cu urmărire sursă
2. "Centralizator" - Articole deduplicate cu cantități agregate și nivele de stoc

Aplicația gestionează documente de business românești cu detecție inteligentă de headere,
validare de date, integrare stoc prin potrivire fuzzy și păstrează formatarea originală.
Zone de focus cheie: flexibilitate mapare headere, reguli de validare date, algoritmi de potrivire stoc,
și formatare output Excel. Vezi DOCUMENTATION.md pentru detalii tehnice complete.

PREZENTARE GENERALĂ ARHITECTURĂ:
===============================
- ExcelProcessorGUI: UI bazat pe Tkinter cu suport threading
- ExcelProcessor: Logică de business centrală cu pipeline de procesare modular
- Metode cheie: process_files(), extract_data_from_sheet(), create_centralizator_data()
- Flux de date: Încărcare fișiere → Detecție headere → Extragere date → Integrare stoc → 
  Consolidare → Generare Centralizator → Output Excel cu păstrare formatare

GHID DE MODIFICARE:
===================
- Menține compatibilitatea înapoi cu utilizatorii existenți
- Testează temeinic cu fișiere eșantion (în special "Oferta Consolight_M.xlsx")
- Extinde funcționalitatea incremental în loc să rescrii logica de bază
- Actualizează DOCUMENTATION.md când adaugi funcționalități noi
- Concentrează-te pe îmbunătățirea mapării headerelor, regulilor de validare și potrivirii stocului

VERSIUNE: 1.1
ULTIMA ACTUALIZARE: August 2025

Schimbări față de 1.0:
- Actualizări UI thread-safe prin Queue + root.after
- Progress determinate cu suport anulare
- Prompt suprasciere și afișare cale completă
- Parsing robust numere EU; agregare corectă cu PU ponderat
- Potrivire fuzzy stoc mai rapidă, indexată
- Copiază celule îmbinate, freeze panes, auto-filter
- Fără handle-uri workbook stagnante; cleanup try/finally
- Mapare headere mai sigură; "RON" opțional în headere
- Blochează selecția .xls legacy

"""

import sys
import os
import re
import threading
import traceback
from typing import List, Dict, Any, Optional, Tuple
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from queue import Queue, Empty

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --------------------------- Utilities ---------------------------

def to_float(x: Any) -> Optional[float]:
    """Parse numbers from EU and US formats, return None on failure."""
    if x is None or x == "":
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace("\xa0", " ")
    s = re.sub(r"\s+", "", s)
    # EU decimal comma
    if re.search(r"\d+,\d+$", s):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None

class ExcelProcessingCancelled(Exception):
    pass

# --------------------------- GUI ---------------------------

class ExcelProcessorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Sheet Consolidator")
        self.root.geometry("900x650")
        self.root.configure(bg='#f0f0f0')

        self.style = ttk.Style()
        self.style.theme_use('clam')

        self.selected_files: List[str] = []
        self.processing = False
        self.cancel_event = threading.Event()

        # single UI queue for status/progress
        self.ui_queue: Queue = Queue()

        self._setup_ui()
        self._bind_shortcuts()
        self.root.after(100, self._poll_ui_queue)

    def _setup_ui(self):
        main = ttk.Frame(self.root, padding="20")
        main.grid(row=0, column=0, sticky=tk.NSEW)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main.columnconfigure(0, weight=1)

        title = ttk.Label(main, text="Excel Sheet Consolidator", font=('Arial', 16, 'bold'))
        title.grid(row=0, column=0, sticky=tk.W, pady=(0, 10))

        instructions = ttk.Label(
            main,
            text="Add .xlsx/.xlsm files. All sheets will be consolidated into 'Cumulative' and 'Centralizator'.",
            font=('Arial', 10)
        )
        instructions.grid(row=1, column=0, sticky=tk.W, pady=(0, 10))

        lf = ttk.LabelFrame(main, text="Selected Files (full paths)", padding="10")
        lf.grid(row=2, column=0, sticky=tk.NSEW)
        lf.columnconfigure(0, weight=1)
        lf.rowconfigure(1, weight=1)

        # Listbox with scrollbars
        list_frame = ttk.Frame(lf)
        list_frame.grid(row=1, column=0, sticky=tk.NSEW)
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)

        self.file_listbox = tk.Listbox(list_frame, height=10, bg='white', selectmode=tk.EXTENDED)
        self.file_listbox.grid(row=0, column=0, sticky=tk.NSEW)

        yscroll = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.file_listbox.yview)
        yscroll.grid(row=0, column=1, sticky=tk.NS)
        self.file_listbox.configure(yscrollcommand=yscroll.set)

        xscroll = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL, command=self.file_listbox.xview)
        xscroll.grid(row=1, column=0, columnspan=2, sticky=tk.EW)
        self.file_listbox.configure(xscrollcommand=xscroll.set)

        btn_row = ttk.Frame(lf)
        btn_row.grid(row=2, column=0, pady=(10, 0), sticky=tk.W)

        self.add_files_btn = ttk.Button(btn_row, text="Add Files", command=self._add_files)
        self.add_files_btn.grid(row=0, column=0, padx=(0, 10))

        self.remove_files_btn = ttk.Button(btn_row, text="Remove Selected", command=self._remove_selected_files)
        self.remove_files_btn.grid(row=0, column=1, padx=(0, 10))

        self.clear_files_btn = ttk.Button(btn_row, text="Clear All", command=self._clear_files)
        self.clear_files_btn.grid(row=0, column=2)

        # Processing controls
        proc_row = ttk.Frame(main)
        proc_row.grid(row=3, column=0, pady=(15, 10), sticky=tk.W)

        self.process_btn = ttk.Button(proc_row, text="Process Files", command=self._process_files)
        self.process_btn.grid(row=0, column=0, padx=(0, 10))

        self.cancel_btn = ttk.Button(proc_row, text="Cancel", command=self._cancel_processing, state="disabled")
        self.cancel_btn.grid(row=0, column=1)

        # Determinate progress bar
        self.progress = ttk.Progressbar(main, length=500, mode='determinate', maximum=100, value=0)
        self.progress.grid(row=4, column=0, sticky=tk.W, pady=(0, 10))

        # Status
        self.status_label = ttk.Label(main, text="Ready", font=('Arial', 10))
        self.status_label.grid(row=5, column=0, sticky=tk.W)

        # expand rows
        main.rowconfigure(2, weight=1)

    def _bind_shortcuts(self):
        self.root.bind("<Control-o>", lambda e: self._add_files())
        self.root.bind("<Control-O>", lambda e: self._add_files())

    # ---------------- File list management ----------------

    def _add_files(self):
        files = filedialog.askopenfilenames(
            title="Select Excel Files",
            filetypes=[("Excel files", "*.xlsx *.xlsm")]
        )
        added = 0
        valid_ext = {".xlsx", ".xlsm"}
        for path in files:
            ext = os.path.splitext(path)[1].lower()
            if ext not in valid_ext:
                messagebox.showwarning("Unsupported file", f"Skipping (unsupported): {os.path.basename(path)}")
                continue
            if path not in self.selected_files:
                self.selected_files.append(path)
                self.file_listbox.insert(tk.END, path)  # full path display
                added += 1
        if added:
            self._set_status(f"Added {added} file(s).")

    def _remove_selected_files(self):
        idxs = list(self.file_listbox.curselection())
        idxs.reverse()
        for i in idxs:
            self.file_listbox.delete(i)
            del self.selected_files[i]
        if idxs:
            self._set_status("Removed selected files.")

    def _clear_files(self):
        self.file_listbox.delete(0, tk.END)
        self.selected_files.clear()
        self._set_status("Cleared all files.")

    # ---------------- Processing flow ----------------

    def _process_files(self):
        if not self.selected_files:
            messagebox.showwarning("No Files", "Add some Excel files first.")
            return

        output_file = filedialog.asksaveasfilename(
            title="Save Consolidated File As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not output_file:
            return

        # Prompt before overwrite
        if os.path.exists(output_file):
            ok = messagebox.askyesno("Overwrite?",
                                     f"'{output_file}' exists. Overwrite?",
                                     icon="warning", default="no")
            if not ok:
                return

        if self.processing:
            return

        # UI state
        self.processing = True
        self.cancel_event.clear()
        self.process_btn.config(text="Processing...", state="disabled")
        self.cancel_btn.config(state="normal")
        self.add_files_btn.config(state="disabled")
        self.remove_files_btn.config(state="disabled")
        self.clear_files_btn.config(state="disabled")
        self.progress.configure(mode='determinate', maximum=100, value=0)
        self._set_status("Starting...")

        # Start worker thread
        thread = threading.Thread(target=self._process_files_thread, args=(output_file,), daemon=True)
        thread.start()

    def _cancel_processing(self):
        if self.processing:
            self.cancel_event.set()
            self._set_status("Cancel requested...")

    def _process_files_thread(self, output_file: str):
        """Worker: never touch Tk variables directly here."""
        try:
            processor = ExcelProcessor(
                cancel_event=self.cancel_event,
                status_cb=lambda m: self.ui_queue.put(("status", m)),
                progress_cb=lambda kind, val: self.ui_queue.put(("progress", kind, val))
            )
            # initialize progress to coarse estimate; will adjust later
            self.ui_queue.put(("progress", "init", 10))

            processor.process_files(self.selected_files, output_file)

            if self.cancel_event.is_set():
                # treated as graceful cancel
                self.root.after(0, self._processing_complete, False, "Cancelled.")
            else:
                self.root.after(0, self._processing_complete, True, "Processing completed.")
        except ExcelProcessingCancelled:
            self.root.after(0, self._processing_complete, False, "Cancelled.")
        except Exception as e:
            err = f"Error during processing: {e}"
            print(err)
            print(traceback.format_exc())
            self.root.after(0, self._processing_complete, False, err)
        finally:
            # ensure UI gets a final tick to 100 if finished
            self.ui_queue.put(("progress", "to_max", None))

    def _processing_complete(self, success: bool, message: str):
        self.processing = False
        self.process_btn.config(text="Process Files", state="normal")
        self.cancel_btn.config(state="disabled")
        self.add_files_btn.config(state="normal")
        self.remove_files_btn.config(state="normal")
        self.clear_files_btn.config(state="normal")
        if success:
            messagebox.showinfo("Success", message)
        else:
            # Avoid showing an error on user cancel
            if "Cancel" in message:
                messagebox.showinfo("Cancelled", message)
            else:
                messagebox.showerror("Error", message)
        self._set_status(message)

    # ---------------- UI plumbing ----------------

    def _set_status(self, msg: str):
        self.status_label.config(text=msg)
        self.root.update_idletasks()

    def _poll_ui_queue(self):
        try:
            while True:
                item = self.ui_queue.get_nowait()
                if not item:
                    break
                kind = item[0]
                if kind == "status":
                    _, msg = item
                    self.status_label.config(text=msg)
                elif kind == "progress":
                    _, action, val = item
                    if action == "init":
                        self.progress.configure(maximum=int(val), value=0)
                    elif action == "add_max":
                        self.progress.configure(maximum=self.progress["maximum"] + int(val))
                    elif action == "inc":
                        new_val = min(self.progress["value"] + int(val), self.progress["maximum"])
                        self.progress.configure(value=new_val)
                    elif action == "set":
                        self.progress.configure(value=int(val))
                    elif action == "to_max":
                        self.progress.configure(value=self.progress["maximum"])
        except Empty:
            pass
        finally:
            self.root.after(100, self._poll_ui_queue)

# --------------------------- Processor ---------------------------

class ExcelProcessor:
    def __init__(self, cancel_event: threading.Event,
                 status_cb, progress_cb):
        self.cancel_event = cancel_event
        self.status = status_cb
        self.progress = progress_cb

        self.common_headers = [
            'Nr. crt', 'Sheet', 'Descriere', 'Denumire', 'Cod articol', 'Furnizor',
            'Cantitate', 'P.U.\n(RON)', 'Pret total\n(RON)',
            'P.U. Taxa\nVerde (RON)', 'Pret total Taxa\nVerde (RON)'
        ]
        self.centralizator_headers = [
            'Descriere', 'Denumire', 'Cod articol', 'Furnizor', 'Cantitate',
            'P.U.\n(RON)', 'Pret total\n(RON)',
            'P.U. Taxa\nVerde (RON)', 'Pret total Taxa\nVerde (RON)', 'Stoc', 'Cod - Stoc'
        ]
        self.source_sheets: List[Dict[str, str]] = []
        self.stock_data: Dict[str, float] = {}
        self._stock_exact: Dict[str, float] = {}
        self._stock_by_suffix: Dict[str, List[Tuple[str, float]]] = {}

    # --------------- Progress helpers ---------------

    def _progress_init(self, max_units: int):
        self.progress("init", max_units)

    def _progress_add_max(self, delta: int):
        self.progress("add_max", delta)

    def _progress_inc(self, step: int = 1):
        self.progress("inc", step)

    def _check_cancel(self):
        if self.cancel_event.is_set():
            raise ExcelProcessingCancelled()

    # --------------- Stock helpers ---------------

    @staticmethod
    def _normalize_code(s: Any) -> str:
        return "".join(str(s).strip().upper().split())

    def _build_stock_index(self):
        self._stock_exact = {self._normalize_code(k): v for k, v in self.stock_data.items()}
        self._stock_by_suffix.clear()
        for k, v in self.stock_data.items():
            nk = self._normalize_code(k)
            suf = nk[-6:] if len(nk) >= 6 else nk
            self._stock_by_suffix.setdefault(suf, []).append((nk, v))

    def find_stock_quantity(self, cod_articol: str) -> float:
        if not cod_articol or not self.stock_data:
            return 0.0
        key = self._normalize_code(cod_articol)
        if key in self._stock_exact:
            return self._stock_exact[key]
        bucket = self._stock_by_suffix.get(key[-6:] if len(key) >= 6 else key, [])
        for nk, v in bucket:
            if nk.endswith(key) or f" {key} " in f" {nk} ":
                return v
        return 0.0

    def find_stock_code(self, cod_articol: str) -> str:
        """Find and return the original stock code that matches the given article code."""
        if not cod_articol or not self.stock_data:
            return ""
        key = self._normalize_code(cod_articol)
        
        # Check exact match first - return the original key from stock_data
        for original_stock_code in self.stock_data.keys():
            if self._normalize_code(original_stock_code) == key:
                return original_stock_code
        
        # Fuzzy matching - return the original stock code that matches
        bucket = self._stock_by_suffix.get(key[-6:] if len(key) >= 6 else key, [])
        for nk, v in bucket:
            if nk.endswith(key) or f" {key} " in f" {nk} ":
                # Find the original stock code that corresponds to this normalized key
                for original_stock_code in self.stock_data.keys():
                    if self._normalize_code(original_stock_code) == nk:
                        return original_stock_code
        return ""

    # --------------- Main pipeline ---------------

    def process_files(self, file_paths: List[str], output_file: str):
        self._check_cancel()
        consolidated_data: List[Dict[str, Any]] = []
        row_counter = 1
        self.source_sheets.clear()

        # Initial progress estimate: per file step + stock + write sheets + copy sheets later
        self._progress_init(max_units=max(10, len(file_paths) + 4))

        try:
            for i, file_path in enumerate(file_paths, 1):
                self._check_cancel()
                self.status(f"Processing file {i}/{len(file_paths)}: {os.path.basename(file_path)}")
                file_data, sheets_info = self._extract_data_from_file(file_path)

                # store sheet refs for later copy (name + path only)
                self.source_sheets.extend(sheets_info)

                # accumulate consolidated rows
                for row in file_data:
                    row['Nr. crt'] = row_counter
                    consolidated_data.append(row)
                    row_counter += 1

                self._progress_inc(1)

            if not consolidated_data:
                raise ValueError("No data was extracted from the files.")

            # Now we know how many sheets to copy; add to progress maximum
            self._progress_add_max(len(self.source_sheets))

            # Stock extraction from first file with suitable sheet
            self._check_cancel()
            self.status("Extracting stock data...")
            try:
                wb_tmp = load_workbook(file_paths[0], data_only=True)
                self.stock_data = self._extract_stock_data(wb_tmp)
                wb_tmp.close()
            except Exception as e:
                self.status(f"Stock extraction warning: {e}")
                self.stock_data = {}

            self._build_stock_index()
            self._progress_inc(1)

            # Create consolidated workbook
            self._check_cancel()
            self.status("Creating consolidated workbook...")
            self._create_consolidated_workbook(consolidated_data, output_file)
            # progress stepped within _create_consolidated_workbook
        except:
            # ensure we propagate after cleanup
            raise

    # --------------- Extraction ---------------

    def _extract_data_from_file(self, file_path: str):
        data: List[Dict[str, Any]] = []
        sheets_info: List[Dict[str, str]] = []
        wb_data = None
        wb_orig = None
        try:
            wb_data = load_workbook(file_path, data_only=True)
            wb_orig = load_workbook(file_path, data_only=False)

            for sheet_name in wb_data.sheetnames:
                self._check_cancel()
                sh = wb_data[sheet_name]
                extracted = self._extract_data_from_sheet(sh, sheet_name)

                for row in extracted:
                    row['Sheet'] = sheet_name
                data.extend(extracted)

                sheets_info.append({'name': sheet_name, 'file_path': file_path})
        except Exception as e:
            raise ValueError(f"Error reading '{os.path.basename(file_path)}': {e}")
        finally:
            # try/finally cleanup
            try:
                if wb_data:
                    wb_data.close()
            finally:
                if wb_orig:
                    wb_orig.close()
        return data, sheets_info

    def _extract_data_from_sheet(self, sheet, sheet_name: str) -> List[Dict[str, Any]]:
        data: List[Dict[str, Any]] = []

        # header row detection
        header_row = None
        for r in range(1, min(25, sheet.max_row + 1)):
            vals = []
            for c in range(1, min(15, sheet.max_column + 1)):
                v = sheet.cell(r, c).value
                if v:
                    vals.append(str(v).strip().lower())
            indicators = ['nr', 'crt', 'descriere', 'denumire', 'cod', 'furnizor', 'cantitate']
            matches = sum(1 for ind in indicators if any(ind in cell for cell in vals))
            if matches >= 3:
                header_row = r
                break

        if not header_row:
            # skip sheet quietly
            return data

        # header mapping
        header_map: Dict[int, str] = {}
        for c in range(1, sheet.max_column + 1):
            raw = sheet.cell(header_row, c).value
            if raw:
                mapped = self._map_header(str(raw))
                if mapped:
                    header_map[c] = mapped

        # extract rows
        extracted = 0
        for r in range(header_row + 1, sheet.max_row + 1):
            if r % 50 == 0:
                self._check_cancel()
            row_data: Dict[str, Any] = {}
            has_data = False
            for c in range(1, sheet.max_column + 1):
                if c in header_map:
                    val = sheet.cell(r, c).value
                    row_data[header_map[c]] = val
                    if val is not None and str(val).strip():
                        has_data = True

            if has_data and self._is_valid_item_row(row_data):
                # ensure all known headers exist (except 'Sheet')
                for h in self.common_headers:
                    if h != 'Sheet' and h not in row_data:
                        row_data[h] = None
                data.append(row_data)
                extracted += 1

        return data

    def _map_header(self, header_text: str) -> Optional[str]:
        h = header_text.lower().strip().replace('\n', ' ')
        if any(p in h for p in ['nr. crt', 'nr crt', 'numar crt', 'nr.crt']):
            return 'Nr. crt'
        if any(p in h for p in ['descriere', 'description']):
            return 'Descriere'
        if any(p in h for p in ['denumire', 'denomination', 'nume']):
            return 'Denumire'
        if any(p in h for p in ['cod articol', 'cod art', 'article code', 'cod produs', 'code']):
            return 'Cod articol'
        if any(p in h for p in ['furnizor', 'supplier', 'provider']):
            return 'Furnizor'
        if any(p in h for p in ['cantitate', 'qty', 'quantity', 'cant']):
            return 'Cantitate'
        # unit price non-green
        if (('p.u.' in h or 'pret unitar' in h or 'unit price' in h)
                and 'taxa' not in h and 'verde' not in h):
            return 'P.U.\n(RON)'
        # total non-green
        if (('pret total' in h or 'total price' in h)
                and 'taxa' not in h and 'verde' not in h):
            return 'Pret total\n(RON)'
        # unit price green
        if (('p.u.' in h or 'pret unitar' in h) and ('taxa' in h or 'verde' in h)):
            return 'P.U. Taxa\nVerde (RON)'
        # total green
        if (('pret total' in h or 'total' in h) and ('taxa' in h or 'verde' in h)):
            return 'Pret total Taxa\nVerde (RON)'
        return None

    def _is_total_row(self, row: Dict[str, Any]) -> bool:
        d = row.get('Descriere', '')
        if isinstance(d, str):
            dl = d.lower().strip()
            for k in ['total', 'suma', 'subtotal', 'total materiale', 'total general', 'sumă', 'sumar', 'consolidat']:
                if k in dl or dl == k:
                    return True
        return False

    def _is_header_row(self, row: Dict[str, Any]) -> bool:
        descr = str(row.get('Descriere', '')).lower().strip()
        if any(k in descr for k in ['descriere', 'denumire', 'description']):
            return True
        other = [str(row.get(x, '')).lower().strip() for x in ['Denumire', 'Cod articol', 'Furnizor']]
        patterns = ['denumire', 'cod articol', 'furnizor', 'cantitate', 'p.u.', 'pret total']
        matches = sum(1 for val in [descr] + other if any(p in val for p in patterns))
        return matches >= 2

    def _is_valid_item_row(self, row: Dict[str, Any]) -> bool:
        descr = row.get('Descriere', '')
        if not descr or not str(descr).strip():
            return False
        if self._is_header_row(row):
            return False
        if self._is_total_row(row):
            return False
        # require at least 2 meaningful fields
        fields = ['Descriere', 'Denumire', 'Cod articol', 'Furnizor', 'Cantitate']
        count = 0
        for f in fields:
            v = row.get(f, '')
            if v is not None and str(v).strip() and str(v).strip().lower() not in ['none', 'n/a', '-']:
                count += 1
        return count >= 2

    # --------------- Stock ---------------

    def _extract_stock_data(self, workbook) -> Dict[str, float]:
        stock: Dict[str, float] = {}

        try:
            # case-insensitive sheet match
            stoc_name = next((n for n in workbook.sheetnames if str(n).lower() in {"stoc", "stock", "inventory"}), None)
            if not stoc_name:
                self.status("No 'Stoc' sheet found. Stock defaults to 0.")
                return {}
            ws = workbook[stoc_name]

            # find header row
            header_row = None
            for r in range(1, min(15, ws.max_row + 1)):
                row_text = " ".join(
                    str(ws.cell(r, c).value).lower()
                    for c in range(1, min(15, ws.max_column + 1))
                    if ws.cell(r, c).value is not None
                )
                if any(k in row_text for k in ['cod', 'stoc', 'total', 'valoare']):
                    header_row = r
                    break
            if not header_row:
                return {}

            cod_col = None
            stoc_col = None
            for c in range(1, ws.max_column + 1):
                v = str(ws.cell(header_row, c).value or '').lower()
                if 'cod' in v:
                    cod_col = c
                if 'stoc' in v or 'total final' in v:
                    stoc_col = c

            if not cod_col or not stoc_col:
                # fallback guess: first col code, first numeric col is stock
                cod_col = cod_col or 1
                for c in range(2, min(10, ws.max_column + 1)):
                    sample = ws.cell(header_row + 1, c).value
                    if sample is not None:
                        try:
                            float(sample)
                            stoc_col = c
                            break
                        except (ValueError, TypeError):
                            pass

            if not cod_col or not stoc_col:
                return {}

            for r in range(header_row + 1, ws.max_row + 1):
                code = ws.cell(r, cod_col).value
                qty = ws.cell(r, stoc_col).value
                if code and qty is not None:
                    fqty = to_float(qty)
                    if fqty is not None:
                        stock[str(code).strip()] = fqty
        except Exception as e:
            self.status(f"Stock extraction error: {e}")
            return {}
        return stock

    # --------------- Aggregation ---------------

    def _create_centralizator_data(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        from collections import defaultdict
        grouped = defaultdict(list)
        items_without_code_grouped = defaultdict(list)  # Group items without codes by description
        
        for it in rows:
            code = str(it.get('Cod articol', '') or '').strip()
            if code:
                grouped[code].append(it)
            else:
                # Group items without code by their description/name
                desc = str(it.get('Descriere', '') or '').strip()
                if desc:
                    items_without_code_grouped[desc].append(it)
                else:
                    # If no description either, use denomination
                    denom = str(it.get('Denumire', '') or '').strip()
                    key = denom if denom else "NO_IDENTIFIER"
                    items_without_code_grouped[key].append(it)

        out: List[Dict[str, Any]] = []
        
        # Helper function for weighted average
        def wavg(pairs: List[Tuple[float, float]]) -> float:
            total_q = sum(q for _, q in pairs)
            return round(sum(p * q for p, q in pairs) / total_q, 6) if total_q else 0.0
        
        # Process grouped items (with codes)
        for code, items in grouped.items():
            sum_qty = 0.0
            sum_total = 0.0
            sum_tax_total = 0.0
            pu_pairs: List[Tuple[float, float]] = []
            tpu_pairs: List[Tuple[float, float]] = []

            base = items[0]

            for it in items:
                q = to_float(it.get('Cantitate')) or 0.0
                pu = to_float(it.get('P.U.\n(RON)'))
                tot = to_float(it.get('Pret total\n(RON)'))
                tpu = to_float(it.get('P.U. Taxa\nVerde (RON)'))
                ttot = to_float(it.get('Pret total Taxa\nVerde (RON)'))

                sum_qty += q
                sum_total += (tot if tot is not None else ((pu or 0.0) * q))
                sum_tax_total += (ttot if ttot is not None else ((tpu or 0.0) * q))
                if pu is not None:
                    pu_pairs.append((pu, q))
                if tpu is not None:
                    tpu_pairs.append((tpu, q))

            p_u = wavg(pu_pairs)
            p_u_taxa = wavg(tpu_pairs)

            # convert qty to int if whole
            qty_val = int(sum_qty) if abs(sum_qty - int(sum_qty)) < 1e-9 else round(sum_qty, 6)

            stock_qty = self.find_stock_quantity(code)
            stock_code = self.find_stock_code(code)

            out.append({
                'Descriere': base.get('Descriere', ''),
                'Denumire': base.get('Denumire', ''),
                'Cod articol': code,
                'Furnizor': base.get('Furnizor', ''),
                'Cantitate': qty_val,
                'P.U.\n(RON)': p_u,
                'Pret total\n(RON)': round(sum_total, 6),
                'P.U. Taxa\nVerde (RON)': p_u_taxa,
                'Pret total Taxa\nVerde (RON)': round(sum_tax_total, 6),
                'Stoc': stock_qty,
                'Cod - Stoc': stock_code
            })
        
        # Process grouped items without codes (aggregated by description/name)
        for desc_key, items in items_without_code_grouped.items():
            sum_qty = 0.0
            sum_total = 0.0
            sum_tax_total = 0.0
            pu_pairs: List[Tuple[float, float]] = []
            tpu_pairs: List[Tuple[float, float]] = []

            base = items[0]

            for it in items:
                q = to_float(it.get('Cantitate')) or 0.0
                pu = to_float(it.get('P.U.\n(RON)'))
                tot = to_float(it.get('Pret total\n(RON)'))
                tpu = to_float(it.get('P.U. Taxa\nVerde (RON)'))
                ttot = to_float(it.get('Pret total Taxa\nVerde (RON)'))

                sum_qty += q
                sum_total += (tot if tot is not None else ((pu or 0.0) * q))
                sum_tax_total += (ttot if ttot is not None else ((tpu or 0.0) * q))
                if pu is not None:
                    pu_pairs.append((pu, q))
                if tpu is not None:
                    tpu_pairs.append((tpu, q))

            p_u = wavg(pu_pairs)
            p_u_taxa = wavg(tpu_pairs)

            # convert qty to int if whole
            qty_val = int(sum_qty) if abs(sum_qty - int(sum_qty)) < 1e-9 else round(sum_qty, 6)

            out.append({
                'Descriere': base.get('Descriere', ''),
                'Denumire': base.get('Denumire', ''),
                'Cod articol': '',  # Empty since no code
                'Furnizor': base.get('Furnizor', ''),
                'Cantitate': qty_val,
                'P.U.\n(RON)': p_u,
                'Pret total\n(RON)': round(sum_total, 6),
                'P.U. Taxa\nVerde (RON)': p_u_taxa,
                'Pret total Taxa\nVerde (RON)': round(sum_tax_total, 6),
                'Stoc': 0.0,  # No stock match possible without code
                'Cod - Stoc': ''  # No stock code match possible without code
            })

        out.sort(key=lambda x: (x.get('Cod articol', ''), x.get('Descriere', '')))
        return out

    # --------------- Workbook creation ---------------

    def _create_consolidated_workbook(self, data: List[Dict[str, Any]], output_file: str):
        from copy import copy

        self._check_cancel()
        self.status("Writing sheets...")

        wb = Workbook()
        try:
            # Cumulative
            ws_cum = wb.active
            ws_cum.title = "Cumulative"
            self._write_cumulative(ws_cum, data)
            self._progress_inc(1)

            # Centralizator
            ws_cen = wb.create_sheet(title="Centralizator", index=1)
            cen_data = self._create_centralizator_data(data)
            self._write_centralizator(ws_cen, cen_data)
            self._progress_inc(1)

            # Copy original sheets
            for i, sh in enumerate(self.source_sheets, 1):
                self._check_cancel()
                src_path = sh['file_path']
                src_name = sh['name']

                # ensure unique name
                base_name = src_name
                unique = base_name
                idx = 1
                while unique in wb.sheetnames:
                    idx += 1
                    unique = f"{base_name}_{idx}"

                ws_new = wb.create_sheet(title=unique)
                # reopen source workbook on demand
                wbs = load_workbook(src_path, data_only=False)
                try:
                    src_ws = wbs[src_name]
                    self._copy_sheet_content(src_ws, ws_new)
                finally:
                    wbs.close()

                self._progress_inc(1)
                self.status(f"Copied sheet {i}/{len(self.source_sheets)}: {src_name}")

            # Save
            self._check_cancel()
            try:
                wb.save(output_file)
            except PermissionError as e:
                raise PermissionError(f"Cannot write '{output_file}': {e}")
        finally:
            wb.close()

    def _write_cumulative(self, ws, rows: List[Dict[str, Any]]):
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # headers
        for c, h in enumerate(self.common_headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_thin
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid") if h == 'Sheet' else header_fill

        # rows
        for r_idx, row in enumerate(rows, 2):
            for c, h in enumerate(self.common_headers, 1):
                val = row.get(h)
                # compute missing totals if possible
                if h == 'Pret total\n(RON)':
                    pt = to_float(val)
                    if pt is None:
                        q = to_float(row.get('Cantitate'))
                        pu = to_float(row.get('P.U.\n(RON)'))
                        if q is not None and pu is not None:
                            val = q * pu
                elif h == 'Pret total Taxa\nVerde (RON)':
                    pt = to_float(val)
                    if pt is None:
                        q = to_float(row.get('Cantitate'))
                        pu = to_float(row.get('P.U. Taxa\nVerde (RON)'))
                        if q is not None and pu is not None:
                            val = q * pu
                cell = ws.cell(row=r_idx, column=c, value=val)
                cell.border = border_thin
                if h in ['Cantitate']:
                    if isinstance(val, (int, float)):
                        cell.number_format = '#,##0'
                if h in ['P.U.\n(RON)', 'Pret total\n(RON)', 'P.U. Taxa\nVerde (RON)', 'Pret total Taxa\nVerde (RON)']:
                    if isinstance(val, (int, float)) and val != 0:
                        cell.number_format = '#,##0.00'

        # autosize
        for c in range(1, len(self.common_headers) + 1):
            letter = get_column_letter(c)
            max_len = 10
            for r in range(1, len(rows) + 2):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[letter].width = min(max_len + 2, 30)

    def _write_centralizator(self, ws, rows: List[Dict[str, Any]]):
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # headers
        for c, h in enumerate(self.centralizator_headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_thin

        # data
        for r_idx, row in enumerate(rows, 2):
            for c, h in enumerate(self.centralizator_headers, 1):
                val = row.get(h)
                cell = ws.cell(row=r_idx, column=c, value=val)
                cell.border = border_thin
                if h in ['Cantitate', 'Stoc']:
                    if isinstance(val, (int, float)):
                        cell.number_format = '#,##0'
                if h in ['P.U.\n(RON)', 'Pret total\n(RON)', 'P.U. Taxa\nVerde (RON)', 'Pret total Taxa\nVerde (RON)']:
                    if isinstance(val, (int, float)) and val != 0:
                        cell.number_format = '#,##0.00'
                if h == 'Cod articol':
                    cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
                if h == 'Cod - Stoc':
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

        # autosize
        for c in range(1, len(self.centralizator_headers) + 1):
            letter = get_column_letter(c)
            max_len = 10
            for r in range(1, len(rows) + 2):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[letter].width = min(max_len + 2, 35)

    def _copy_sheet_content(self, src_ws, dst_ws):
        from copy import copy

        for row in src_ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    nc = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        nc.font = copy(cell.font)
                        nc.fill = copy(cell.fill)
                        nc.border = copy(cell.border)
                        nc.alignment = copy(cell.alignment)
                        nc.number_format = cell.number_format

        # merged ranges
        for mr in src_ws.merged_cells.ranges:
            dst_ws.merge_cells(str(mr))

        # dimensions
        for col_letter, dim in src_ws.column_dimensions.items():
            dst_ws.column_dimensions[col_letter].width = dim.width
        for row_num, dim in src_ws.row_dimensions.items():
            dst_ws.row_dimensions[row_num].height = dim.height

        # panes and filters
        dst_ws.freeze_panes = src_ws.freeze_panes
        if src_ws.auto_filter and src_ws.auto_filter.ref:
            dst_ws.auto_filter.ref = src_ws.auto_filter.ref

# --------------------------- Entrypoint ---------------------------

def main():
    root = tk.Tk()
    app = ExcelProcessorGUI(root)

    # Center window
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (root.winfo_width() // 2)
    y = (root.winfo_screenheight() // 2) - (root.winfo_height() // 2)
    root.geometry(f"+{x}+{y}")

    try:
        root.mainloop()
    except KeyboardInterrupt:
        print("\nInterrupted.")
        sys.exit(0)

if __name__ == "__main__":
    main()
